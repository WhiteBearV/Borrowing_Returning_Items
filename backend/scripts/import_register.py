"""
สร้าง seed_equipment.sql จากไฟล์ทะเบียนคุมทรัพย์สินของคณะ (Excel) — ใช้เลขครุภัณฑ์จริง

ต่างจาก import_equipment.py (ของเดิม ดึงจาก Snipe-IT ซึ่งเลขไม่ตรงทะเบียน) — ตัวนี้อ่าน
ไฟล์ทะเบียนตรง ๆ เป็น source of truth เดียว:
  - ชีต "คณะเทคโนฯดิจิทัล"        → code = คอลัมน์ "รหัสครุภัณฑ์"
  - ชีต "ครุภัณฑ์ที่ได้รับพระราชทาน" → code = คอลัมน์ "รหัสครุภัณฑ์ใหม่" (ตามที่อาจารย์สั่ง)

Usage:
    pip install openpyxl
    python backend/scripts/import_register.py

Output:
    backend/scripts/seed_equipment.sql   — SQL seed (idempotent, UUID เสถียรด้วย uuid5)
    Report/ครุภัณฑ์ไม่มีเลข.md            — รายการที่ไม่มีเลขครุภัณฑ์ (ให้เจ้าของตัดสินใจ)
"""
import os
import uuid

import openpyxl

# path อิงจาก repo root ไม่ว่าจะรันที่ไหน
ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX = os.path.join(
    ROOT, "docs",
    "รายงานทะเบียนคุมทรัพย์สิน - คณะเทคโนโลยีดิจิทัล ข้อมูล ณ วันที่ 4 พฤศจิกายน 2568.xlsx",
)
SQL_OUT = os.path.join(ROOT, "backend", "scripts", "seed_equipment.sql")
NOCODE_OUT = os.path.join(ROOT, "Report", "ครุภัณฑ์ไม่มีเลข.md")

# namespace คงที่ → uuid5 ให้ id เดิมทุกครั้งที่ gen ใหม่ (idempotent, re-seed ได้)
NS = uuid.UUID("6f1e7c9a-0000-5000-8000-000000000000")

# --- กฎจัดหมวดตามคีย์เวิร์ดในชื่อ (เช็คจากบนลงล่าง เจอก่อนชนะ) ---
# วางคอมพิวเตอร์/ซอฟต์แวร์/AV/อิเล็กทรอนิกส์ ก่อนเฟอร์นิเจอร์
# เพื่อไม่ให้ "คอมพิวเตอร์ตั้งโต๊ะ" หลุดไปเข้าหมวดโต๊ะเพราะมีคำว่า "โต๊ะ"
CATEGORY_RULES: list[tuple[str, list[str]]] = [
    ("คอมพิวเตอร์", ["คอมพิวเตอร์", "optiplex", "acer", "โน๊ตบุ๊ค", "โน้ตบุ๊ค",
                     "thinkpad", "แม่ข่าย", "micro pc", "กราฟฟิก", "หน้าจอ", "จอแสดงผล"]),
    ("ซอฟต์แวร์/ลิขสิทธิ์", ["ซอฟต์แวร์", "ลิขสิทธิ์", "ชุดโปรแกรม"]),
    ("โสตทัศนูปกรณ์", ["โทรทัศน์", "led tv", "smart tv"]),
    ("อุปกรณ์อิเล็กทรอนิกส์/เครื่องมือวัด",
     ["มัลติมิเตอร์", "oscilloscope", "เครื่องวัด", "เครื่องกำเนิด", "แหล่งจ่าย",
      "เครื่องทดสอบ", "ลอจิก", "iot", "ไอโอที", "sensor", "แผงวงจร", "สมองกล",
      "virtual reality", "คลัสเตอร์", "big data", "ไฟฟ้า", "เครื่องมือวัด",
      "อิเล็กทรอนิกส์", "ทดลอง", "ทดสอบ", "อุปกรณ์วัด"]),
    ("โต๊ะ/ตู้/เฟอร์นิเจอร์", ["โต๊ะ", "โตะ", "เก้าอี้", "ตู้", "ชั้นวาง"]),
]
FALLBACK_CAT = "ไม่ระบุหมวดหมู่"


def categorize(name: str) -> str:
    low = name.lower()
    for cat, keys in CATEGORY_RULES:
        if any(k in low for k in keys):
            return cat
    return FALLBACK_CAT


def det_uuid(*parts: str) -> str:
    return str(uuid.uuid5(NS, "|".join(parts)))


def q(s: object) -> str:
    if s is None or s == "":
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def clean(v: object) -> str:
    """ตัด whitespace/ขึ้นบรรทัดที่ติดมาจากเซลล์ Excel"""
    return " ".join(str(v).split()) if v is not None else ""


def money(v: object) -> str:
    """ราคาจากทะเบียน → ค่า SQL — ช่องว่าง/ขีด/ข้อความที่ไม่ใช่ตัวเลข = NULL (ไปกรอกเองทีหลังได้)

    ทะเบียนพระราชทานเก็บทศนิยมยาว (7383.177570093458) — ปัดเป็น 2 ตำแหน่งตามคอลัมน์ Numeric(12,2)
    """
    try:
        return f"{round(float(str(v).replace(',', '')), 2)}"
    except (TypeError, ValueError):
        return "NULL"


def status_from(normal, broken, worn, lost) -> tuple[str, int]:
    """แปลงคอลัมน์ ปกติ/ชำรุด/เสื่อมสภาพ/สูญหาย → (status, qty_available)"""
    if broken:
        return "damaged", 0
    if lost or worn:
        return "unavailable", 0
    return "available", 1


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # (code, name, item_type, status, qty_avail, location, category, unit_value, unit)
    records: list[tuple] = []
    no_code: list[dict] = []

    # --- ชีต 1: คณะ ---
    ws = wb["คณะเทคโนฯดิจิทัล"]
    for r in ws.iter_rows(min_row=5, values_only=True):
        name = clean(r[2])
        if not name or name == "รวมทั้งสิ้น":
            continue
        code = clean(r[1])
        loc = clean(r[11]) or clean(r[16])
        status, qa = status_from(r[5], r[6], r[7], r[8])
        if not code:
            no_code.append({"sheet": "คณะ", "seq": clean(r[0]), "name": name, "loc": loc})
            continue
        # คอลัมน์ 3 = "ราคา" (ชีตนี้ไม่มีหน่วยนับ ครุภัณฑ์นับเป็นชิ้นอยู่แล้ว)
        records.append((code, name, "durable", status, qa, loc, categorize(name), money(r[3]), None))

    # --- ชีต 2: พระราชทาน (ใช้ "รหัสครุภัณฑ์ใหม่" = คอลัมน์ 8) ---
    ws = wb["ครุภัณฑ์ที่ได้รับพระราชทาน"]
    for r in ws.iter_rows(min_row=5, values_only=True):
        name = clean(r[1])
        if not name or name == "รวมทั้งสิ้น":
            continue
        code = clean(r[8])
        if code == "-":
            code = ""
        loc = clean(r[9])
        note = clean(r[15])
        status, qa = status_from(r[10], r[11], r[12], r[13])
        # "ไม่พบครุภัณฑ์" = หาตัวของจริงไม่เจอ → ถือว่ายืมไม่ได้
        if "ไม่พบครุภัณฑ์" in note:
            status, qa = "unavailable", 0
        if not code:
            no_code.append({"sheet": "พระราชทาน", "seq": clean(r[0]), "name": name, "loc": loc})
            continue
        # คอลัมน์ 5 = "ราคา/หน่วย", 4 = "หน่วยนับ" (ตัว/ชุด/เครื่อง) — ใบยืมเอาไปแสดงคู่กับจำนวน
        records.append((code, name, "durable", status, qa, loc, categorize(name),
                        money(r[5]), clean(r[4]) or None))

    # หมวดทั้งหมดที่ใช้จริง
    cats = sorted({rec[6] for rec in records})

    # --- เขียน SQL ---
    lines: list[str] = []
    lines.append("-- ===========================================")
    lines.append("-- Equipment seed จากทะเบียนคุมทรัพย์สินคณะ (ข้อมูล 4 พ.ย. 2568)")
    lines.append("-- gen โดย backend/scripts/import_register.py — ใช้เลขครุภัณฑ์จริง")
    lines.append("-- ===========================================")
    lines.append("BEGIN;")
    lines.append("")
    lines.append("-- Categories")
    for c in cats:
        lines.append(
            f"INSERT INTO equipment_categories (id, name) "
            f"VALUES ({q(det_uuid('cat', c))}, {q(c)}) ON CONFLICT (name) DO NOTHING;"
        )
    lines.append("")
    lines.append(f"-- Equipment ({len(records)} records)")
    for code, name, itype, status, qa, loc, cat, value, unit in records:
        eid = det_uuid("eq", code)
        # ON CONFLICT: เติมราคา/หน่วยให้แถวที่มีอยู่แล้ว แต่ COALESCE ของเดิมไว้ก่อน
        # → ค่าที่แอดมินกรอกเองในหน้าจัดการอุปกรณ์จะไม่ถูก re-seed ทับ (ทะเบียนเป็นแค่ค่าตั้งต้น)
        lines.append(
            "INSERT INTO equipment (id, code, name, item_type, status, "
            "quantity_total, quantity_available, description, location, image_urls, unit_value, unit) VALUES ("
            f"{q(eid)}, {q(code)}, {q(name)}, {q(itype)}, {q(status)}, "
            f"1, {qa}, NULL, {q(loc)}, '[]', {value}, {q(unit)}) "
            "ON CONFLICT (code) DO UPDATE SET "
            "unit_value = COALESCE(equipment.unit_value, EXCLUDED.unit_value), "
            "unit = COALESCE(equipment.unit, EXCLUDED.unit);"
        )
        lines.append(
            "INSERT INTO equipment_category_links (equipment_id, category_id) "
            f"SELECT id, {q(det_uuid('cat', cat))} FROM equipment WHERE code = {q(code)} "
            "ON CONFLICT DO NOTHING;"
        )
    lines.append("")
    lines.append("COMMIT;")
    with open(SQL_OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    # --- เขียนรายการไม่มีเลขครุภัณฑ์ ---
    md = ["# ครุภัณฑ์ที่ไม่มีเลขในทะเบียน (รอเจ้าของตัดสินใจ)", "",
          f"พบ {len(no_code)} รายการที่ไม่มีเลขครุภัณฑ์ในไฟล์ทะเบียน — ยังไม่ถูกใส่ลง seed",
          "รอตัดสินใจว่าจะออกเลขภายในเอง / ข้าม / หรือรอเลขจากงานพัสดุ", "",
          "| ชีต | ลำดับ | รายการ | สถานที่ |", "|---|---|---|---|"]
    for n in no_code:
        md.append(f"| {n['sheet']} | {n['seq']} | {n['name']} | {n['loc']} |")
    with open(NOCODE_OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(md) + "\n")

    # สรุปลง stdout
    print(f"equipment: {len(records)} rows, categories: {len(cats)}")
    from collections import Counter
    for c, n in Counter(rec[6] for rec in records).most_common():
        print(f"  {n:4}  {c}")
    print(f"ไม่มีเลขครุภัณฑ์: {len(no_code)} rows -> {NOCODE_OUT}")
    print(f"SQL -> {SQL_OUT}")


if __name__ == "__main__":
    main()
