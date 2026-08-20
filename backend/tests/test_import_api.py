"""นำเข้าจาก Excel ผ่าน API จริง: อัปโหลด → ร่าง → แอดมินแก้/ตัดออก → บันทึกเฉพาะที่เลือก

ตรวจว่าบรรทัดที่แอดมินตัดออกจากร่างต้องไม่เข้าระบบ, ค่าที่แก้ในร่างต้องถูกใช้จริง,
และร่างที่ถูกยัดของแปลกปลอม (code ไม่มีในไฟล์) ต้องถูกปฏิเสธ
"""
import openpyxl
import pytest
import pytest_asyncio
from sqlalchemy import delete, select

from app.core.database import AsyncSessionLocal
from app.models.audit_log import AuditLog
from app.models.equipment import Equipment
from tests.conftest import auth

CODE_A = "TESTIMP-001"
CODE_B = "TESTIMP-002"


@pytest_asyncio.fixture
async def cleanup_imported():
    yield
    async with AsyncSessionLocal() as db:
        ids = (await db.execute(
            select(Equipment.id).where(Equipment.code.in_([CODE_A, CODE_B]))
        )).scalars().all()
        for eid in ids:
            await db.execute(delete(AuditLog).where(AuditLog.target_id == eid))
        await db.execute(delete(Equipment).where(Equipment.code.in_([CODE_A, CODE_B])))
        await db.commit()


def _register_file(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "คณะเทคโนฯดิจิทัล"
    for _ in range(4):
        ws.append([None])
    ws.append([1, CODE_A, "ออสซิลโลสโคป Rigol DS1054Z", 25000, None, "P", None, None, None, None, "คณะ", "15311"])
    ws.append([2, CODE_B, "มัลติมิเตอร์ Fluke 117", 8000, None, "P", None, None, None, None, "คณะ", "15311"])
    path = tmp_path / "register.xlsx"
    wb.save(path)
    return path


@pytest.mark.asyncio(loop_scope="session")
async def test_import_preview_then_commit_selected_rows(client, admin_token, tmp_path, cleanup_imported):
    path = _register_file(tmp_path)
    with open(path, "rb") as f:
        res = await client.post(
            "/equipment/import/preview",
            files={"file": ("register.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth(admin_token),
        )
    assert res.status_code == 200
    draft = res.json()
    new_rows = {r["code"]: r for r in draft["rows"] if r["action"] == "new"}
    assert CODE_A in new_rows and CODE_B in new_rows

    # แอดมินแก้ชื่อ/สถานะในร่าง แล้วตัด CODE_B ออก (ไม่เอาเข้าคลัง)
    res = await client.post(
        f"/equipment/import/{draft['import_id']}/commit",
        json={"filename": "register.xlsx", "rows": [{
            "code": CODE_A, "action": "new", "name": "ออสซิลโลสโคป Rigol (แก้ชื่อในร่าง)",
            "location": "15399", "status": "damaged", "category": "อุปกรณ์อิเล็กทรอนิกส์/เครื่องมือวัด",
        }]},
        headers=auth(admin_token),
    )
    assert res.status_code == 200
    assert res.json()["new"] == 1

    async with AsyncSessionLocal() as db:
        eq = (await db.execute(select(Equipment).where(Equipment.code == CODE_A))).scalar_one()
        assert eq.name == "ออสซิลโลสโคป Rigol (แก้ชื่อในร่าง)"  # ใช้ค่าที่แก้ในร่าง ไม่ใช่ค่าจากไฟล์
        assert eq.location == "15399"
        assert eq.status == "damaged" and eq.quantity_available == 0  # ชำรุด → ยืมไม่ได้
        # บรรทัดที่ถูกตัดออกจากร่างต้องไม่เข้าระบบ
        assert (await db.execute(select(Equipment).where(Equipment.code == CODE_B))).scalar_one_or_none() is None
        # ต้องมี audit log ให้ออกใบรับเข้าคลังได้
        log = (await db.execute(
            select(AuditLog).where(AuditLog.target_id == eq.id, AuditLog.action == "create_equipment")
        )).scalar_one()
        assert log.detail["code"] == CODE_A and log.detail["source"] == "register.xlsx"


@pytest.mark.asyncio(loop_scope="session")
async def test_commit_rejects_row_not_in_uploaded_file(client, admin_token, tmp_path, cleanup_imported):
    path = _register_file(tmp_path)
    with open(path, "rb") as f:
        res = await client.post(
            "/equipment/import/preview",
            files={"file": ("register.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth(admin_token),
        )
    import_id = res.json()["import_id"]

    res = await client.post(
        f"/equipment/import/{import_id}/commit",
        json={"rows": [{"code": "ของแปลกปลอม-999", "action": "new", "name": "ของที่ไม่ได้อยู่ในไฟล์",
                        "status": "available"}]},
        headers=auth(admin_token),
    )
    assert res.status_code == 400


# ── ชีต "SN" เสริม — update-only backfill SN โดยจับคู่กับรหัสที่มีอยู่แล้วในระบบเท่านั้น ─────────────────

CODE_SN_A = "TESTSN-001"
CODE_SN_UNTOUCHED = "TESTSN-002"


@pytest_asyncio.fixture
async def sn_test_equipment():
    """A: จะถูกอัปเดต SN ผ่านไฟล์ · UNTOUCHED: มี SN เดิมอยู่แล้ว ไฟล์ไม่ได้พูดถึง ต้องไม่โดนแตะ"""
    async with AsyncSessionLocal() as db:
        db.add(Equipment(code=CODE_SN_A, name="ทดสอบ SN A", item_type="durable",
                         quantity_total=1, quantity_available=1, status="available", image_urls=[]))
        db.add(Equipment(code=CODE_SN_UNTOUCHED, name="ทดสอบ SN B (ไม่ถูกแตะ)", item_type="durable",
                         quantity_total=1, quantity_available=1, status="available", image_urls=[],
                         serial_number="SN-OLD-UNTOUCHED"))
        await db.commit()
    yield
    async with AsyncSessionLocal() as db:
        ids = (await db.execute(
            select(Equipment.id).where(Equipment.code.in_([CODE_SN_A, CODE_SN_UNTOUCHED]))
        )).scalars().all()
        for eid in ids:
            await db.execute(delete(AuditLog).where(AuditLog.target_id == eid))
        await db.execute(delete(Equipment).where(Equipment.code.in_([CODE_SN_A, CODE_SN_UNTOUCHED])))
        await db.commit()


def _sn_file(tmp_path):
    """ไฟล์ที่มีแต่ชีต "SN" — 1 รหัสจริง (จับคู่ได้) + 1 รหัสมั่ว (ไม่พบในระบบ)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SN"
    ws.append(["รหัส", "SN"])
    ws.append([CODE_SN_A, "SN-NEW-A"])
    ws.append(["รหัสมั่ว-ไม่มีในระบบ-999", "SN-ORPHAN"])
    path = tmp_path / "sn.xlsx"
    wb.save(path)
    return path


@pytest.mark.asyncio(loop_scope="session")
async def test_sn_sheet_preview_flags_unknown_code_as_skipped(client, admin_token, tmp_path, sn_test_equipment):
    path = _sn_file(tmp_path)
    with open(path, "rb") as f:
        res = await client.post(
            "/equipment/import/preview",
            files={"file": ("sn.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth(admin_token),
        )
    assert res.status_code == 200, res.text
    draft = res.json()
    assert any(s["reason"] == "ไม่พบรหัสนี้ในระบบ" for s in draft["skipped"])
    row = next(r for r in draft["rows"] if r["code"] == CODE_SN_A)
    assert row["action"] == "update"
    assert row["changes"]["serial_number"] == [None, "SN-NEW-A"]


@pytest.mark.asyncio(loop_scope="session")
async def test_sn_sheet_commit_updates_matched_code_only(client, admin_token, tmp_path, sn_test_equipment):
    path = _sn_file(tmp_path)
    with open(path, "rb") as f:
        res = await client.post(
            "/equipment/import/preview",
            files={"file": ("sn.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth(admin_token),
        )
    draft = res.json()
    row = next(r for r in draft["rows"] if r["code"] == CODE_SN_A)

    res = await client.post(
        f"/equipment/import/{draft['import_id']}/commit",
        json={"filename": "sn.xlsx", "rows": [{
            "code": row["code"], "action": row["action"], "name": row["name"],
            "location": row["location"], "status": row["status"], "item_type": row["item_type"],
            "quantity": row["quantity_total"], "unit": row["unit"],
            "categories": [], "serial_number": row["serial_number"],
        }]},
        headers=auth(admin_token),
    )
    assert res.status_code == 200, res.text

    async with AsyncSessionLocal() as db:
        a = (await db.execute(select(Equipment).where(Equipment.code == CODE_SN_A))).scalar_one()
        untouched = (await db.execute(select(Equipment).where(Equipment.code == CODE_SN_UNTOUCHED))).scalar_one()
        assert a.serial_number == "SN-NEW-A"
        # regression: แถวที่ไฟล์ไม่ได้พูดถึงต้องไม่ถูกเขียนทับ SN เดิมด้วย null
        assert untouched.serial_number == "SN-OLD-UNTOUCHED"


def _sn_duplicate_file(tmp_path):
    """ไฟล์ชีต SN ที่ 2 รหัสจริงต่างกันอ้าง SN ค่าเดียวกัน — ต้องไม่ผ่านไปเป็น action=update ทั้งคู่"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SN"
    ws.append(["รหัส", "SN"])
    ws.append([CODE_SN_A, "SN-CLASH"])
    ws.append([CODE_SN_UNTOUCHED, "SN-CLASH"])
    path = tmp_path / "sn_dup.xlsx"
    wb.save(path)
    return path


@pytest.mark.asyncio(loop_scope="session")
async def test_sn_sheet_duplicate_value_in_file_is_skipped_not_applied(client, admin_token, tmp_path, sn_test_equipment):
    """สองรหัสอ้าง SN เดียวกันในไฟล์เดียวกัน — ต้องถูกตัดไปที่ skipped ทั้งคู่ ไม่ใช่ action=update ที่จะพัง
    ตอน commit ด้วย UniqueViolationError (ไม่มี IntegrityError handler ไหนใน backend ดักไว้)"""
    path = _sn_duplicate_file(tmp_path)
    with open(path, "rb") as f:
        res = await client.post(
            "/equipment/import/preview",
            files={"file": ("sn_dup.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth(admin_token),
        )
    assert res.status_code == 200, res.text
    draft = res.json()
    skipped_codes = {s["seq"] for s in draft["skipped"] if s["sheet"] == "SN"}
    assert {CODE_SN_A, CODE_SN_UNTOUCHED} <= skipped_codes
    assert not any(r["code"] in (CODE_SN_A, CODE_SN_UNTOUCHED) and "serial_number" in r["changes"]
                   for r in draft["rows"])

    # commit ร่างเปล่า (ไม่มีอะไรให้บันทึกเพราะทั้งคู่โดนตัดไปที่ skipped) ต้องไม่ 500
    res = await client.post(
        f"/equipment/import/{draft['import_id']}/commit",
        json={"filename": "sn_dup.xlsx", "rows": []},
        headers=auth(admin_token),
    )
    assert res.status_code == 200, res.text

    async with AsyncSessionLocal() as db:
        a = (await db.execute(select(Equipment).where(Equipment.code == CODE_SN_A))).scalar_one()
        untouched = (await db.execute(select(Equipment).where(Equipment.code == CODE_SN_UNTOUCHED))).scalar_one()
        assert a.serial_number is None
        assert untouched.serial_number == "SN-OLD-UNTOUCHED"


@pytest.mark.asyncio(loop_scope="session")
async def test_commit_rejects_sn_conflicting_with_different_code_as_409(client, admin_token, tmp_path, sn_test_equipment):
    """defense-in-depth: client ส่ง serial_number ที่ชนกับรหัสอื่นตรง ๆ ใน body (ข้าม validation ตอน preview)
    ต้องได้ 409 อ่านง่าย ไม่ใช่ 500 จาก UniqueViolationError ดิบ ๆ"""
    path = _sn_file(tmp_path)
    with open(path, "rb") as f:
        res = await client.post(
            "/equipment/import/preview",
            files={"file": ("sn.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth(admin_token),
        )
    draft = res.json()
    row = next(r for r in draft["rows"] if r["code"] == CODE_SN_A)

    res = await client.post(
        f"/equipment/import/{draft['import_id']}/commit",
        json={"filename": "sn.xlsx", "rows": [{
            "code": row["code"], "action": row["action"], "name": row["name"],
            "location": row["location"], "status": row["status"], "item_type": row["item_type"],
            "quantity": row["quantity_total"], "unit": row["unit"],
            "categories": [], "serial_number": "SN-OLD-UNTOUCHED",  # ชนกับ CODE_SN_UNTOUCHED ตรง ๆ
        }]},
        headers=auth(admin_token),
    )
    assert res.status_code == 409, res.text

    async with AsyncSessionLocal() as db:
        a = (await db.execute(select(Equipment).where(Equipment.code == CODE_SN_A))).scalar_one()
        assert a.serial_number is None  # ต้องไม่ถูกเขียนเมื่อโดนปฏิเสธ


CODE_NEW_WITH_SN = "TESTIMP-SN-NEW-001"


def _material_and_sn_file(tmp_path):
    """ไฟล์เดียวมีทั้งชีต "วัสดุ" (รหัสใหม่ที่ยังไม่มีใน DB) และชีต "SN" ที่ชี้ไปรหัสเดียวกัน

    regression: เดิม diff_rows ไม่พ่วง serial_number ให้แถว action=new เลย ทำให้รหัสใหม่ที่มี SN คู่กัน
    ในไฟล์เดียวกันถูกสร้างพร้อม SN จริงตอน commit โดยที่แอดมินไม่เคยเห็นค่านี้ในร่างตอน preview มาก่อน
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "วัสดุ"
    ws.append(["รหัส", "รายการ", "ประเภท", "หน่วยนับ", "จำนวน", "สถานที่จัดเก็บ"])
    ws.append([CODE_NEW_WITH_SN, "บอร์ดทดสอบ SN ใหม่", "วัสดุใช้ซ้ำ", "ชิ้น", 1, "15311"])
    ws_sn = wb.create_sheet("SN")
    ws_sn.append(["รหัส", "SN"])
    ws_sn.append([CODE_NEW_WITH_SN, "SN-BRANDNEW-001"])
    path = tmp_path / "material_with_sn.xlsx"
    wb.save(path)
    return path


@pytest.mark.asyncio(loop_scope="session")
async def test_new_row_with_sn_sheet_entry_shows_serial_number_in_preview(client, admin_token, tmp_path):
    path = _material_and_sn_file(tmp_path)
    try:
        with open(path, "rb") as f:
            res = await client.post(
                "/equipment/import/preview",
                files={"file": ("material_with_sn.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                headers=auth(admin_token),
            )
        assert res.status_code == 200, res.text
        draft = res.json()
        row = next(r for r in draft["rows"] if r["code"] == CODE_NEW_WITH_SN)
        assert row["action"] == "new"
        # Issue B: ต้องเห็น SN ตั้งแต่ preview/diff ไม่ใช่แอบถูก apply เงียบ ๆ ตอน commit เท่านั้น
        assert row["serial_number"] == "SN-BRANDNEW-001"
        assert row["changes"].get("serial_number") == [None, "SN-BRANDNEW-001"]

        res = await client.post(
            f"/equipment/import/{draft['import_id']}/commit",
            json={"filename": "material_with_sn.xlsx", "rows": [{
                "code": row["code"], "action": row["action"], "name": row["name"],
                "location": row["location"], "status": row["status"], "item_type": row["item_type"],
                "quantity": row["quantity_total"], "unit": row["unit"],
                "categories": [], "serial_number": row["serial_number"],
            }]},
            headers=auth(admin_token),
        )
        assert res.status_code == 200, res.text

        async with AsyncSessionLocal() as db:
            eq = (await db.execute(select(Equipment).where(Equipment.code == CODE_NEW_WITH_SN))).scalar_one()
            assert eq.serial_number == "SN-BRANDNEW-001"
    finally:
        async with AsyncSessionLocal() as db:
            ids = (await db.execute(
                select(Equipment.id).where(Equipment.code == CODE_NEW_WITH_SN)
            )).scalars().all()
            for eid in ids:
                await db.execute(delete(AuditLog).where(AuditLog.target_id == eid))
            await db.execute(delete(Equipment).where(Equipment.code == CODE_NEW_WITH_SN))
            await db.commit()
