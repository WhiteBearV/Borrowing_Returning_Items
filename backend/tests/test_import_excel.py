"""นำเข้าอุปกรณ์จากไฟล์ทะเบียน Excel — ตรวจว่าอ่านไฟล์แล้ว diff ออกมาถูกต้อง

ครอบคลุมกรณีที่ทำให้ของในคลังเพี้ยนได้: ของใหม่ / สถานะเปลี่ยนเป็นชำรุด / ของที่หายไปจากไฟล์
และวัสดุที่ไม่ได้อยู่ในทะเบียนต้องไม่ถูกเสนอปลดระวาง
"""
import openpyxl
import pytest

from app.services.import_service import _filter_bad_sn_updates, diff_rows, parse_workbook


def _make_register(tmp_path, rows):
    """สร้างไฟล์ทะเบียนจำลอง (ชีตคณะ) — header 4 แถวเหมือนของจริง, ข้อมูลเริ่มแถว 5"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "คณะเทคโนฯดิจิทัล"
    for _ in range(4):
        ws.append([None])
    for i, (code, name, normal, broken, loc) in enumerate(rows, 1):
        # คอลัมน์: 0 ลำดับ, 1 รหัส, 2 ชื่อ, 3 ราคา, 4 วันที่, 5 ปกติ, 6 ชำรุด, 7 เสื่อม, 8 สูญหาย,
        #          9 ผิดประเภท, 10 หน่วยงาน, 11 สถานที่
        ws.append([i, code, name, 1000, None, normal, broken, None, None, None, "คณะ", loc])
    path = tmp_path / "register.xlsx"
    wb.save(path)
    return str(path)


def test_parse_and_diff(tmp_path):
    path = _make_register(tmp_path, [
        ("63-001-0001", "คอมพิวเตอร์ตั้งโต๊ะ Dell", "P", None, "15310"),   # มีอยู่แล้ว ตรงกัน
        ("63-001-0002", "คอมพิวเตอร์ตั้งโต๊ะ Dell", None, "P", "15310"),   # เดิมปกติ → ไฟล์บอกชำรุด
        ("63-001-0003", "จอแสดงผล 27 นิ้ว", "P", None, "15311"),           # ยังไม่มีใน DB
        ("", "โต๊ะไม่มีเลข", "P", None, "5303"),                            # ไม่มีเลข → ข้าม
    ])
    rows, skipped, _sn = parse_workbook(path)
    assert len(rows) == 3
    assert len(skipped) == 1

    existing = {
        "63-001-0001": {"id": 1, "name": "คอมพิวเตอร์ตั้งโต๊ะ Dell", "location": "15310",
                        "status": "available", "item_type": "durable"},
        "63-001-0002": {"id": 2, "name": "คอมพิวเตอร์ตั้งโต๊ะ Dell", "location": "15310",
                        "status": "available", "item_type": "durable"},
        "63-999-9999": {"id": 3, "name": "เครื่องเก่าที่หลุดจากทะเบียน", "location": None,
                        "status": "available", "item_type": "durable"},
        "SNIPE-500": {"id": 4, "name": "สายไฟ", "location": None,
                      "status": "available", "item_type": "consumable"},
    }
    by_code = {d["code"]: d for d in diff_rows(rows, existing)}

    assert by_code["63-001-0001"]["action"] == "unchanged"
    assert by_code["63-001-0002"]["action"] == "update"
    assert by_code["63-001-0002"]["changes"]["status"] == ["available", "damaged"]
    assert by_code["63-001-0003"]["action"] == "new"
    assert by_code["63-001-0003"]["category"] == "คอมพิวเตอร์"  # จัดหมวดจากชื่ออัตโนมัติ
    # ครุภัณฑ์ที่หายจากไฟล์ → เสนอปลดระวาง, แต่วัสดุ (ไม่อยู่ในทะเบียน) ต้องไม่โดน
    assert by_code["63-999-9999"]["action"] == "retire"
    assert "SNIPE-500" not in by_code


def test_parse_material_sheet(tmp_path):
    """ชีต 'วัสดุ' — วัสดุไม่มีเลขครุภัณฑ์ในทะเบียน จึงมาจากชีตแยกพร้อมจำนวน/หน่วย/ประเภท"""
    wb = openpyxl.Workbook()
    wb.active.title = "วัสดุ"
    ws = wb.active
    ws.append(["รหัส", "รายการ", "ประเภท", "หน่วยนับ", "จำนวน", "สถานที่จัดเก็บ"])
    ws.append(["MAT-ARD-001", "บอร์ด Arduino UNO R3", "วัสดุใช้ซ้ำ", "ชิ้น", 25, "15310"])
    ws.append(["CON-WIRE-001", "สายไฟจัมเปอร์", "วัสดุสิ้นเปลือง", "ชุด", 40, "15310"])
    ws.append([None, "ตัวต้านทาน (ยังไม่ออกรหัส)", "วัสดุสิ้นเปลือง", "ถุง", 10, "15310"])
    path = tmp_path / "materials.xlsx"
    wb.save(path)

    rows, skipped, _sn = parse_workbook(str(path))
    assert len(skipped) == 1  # แถวไม่มีรหัสต้องถูกข้าม ไม่ใช่สร้างของไร้รหัส
    by_code = {r["code"]: r for r in rows}
    assert by_code["MAT-ARD-001"]["item_type"] == "material"
    assert by_code["MAT-ARD-001"]["quantity_total"] == 25
    assert by_code["MAT-ARD-001"]["unit"] == "ชิ้น"
    assert by_code["CON-WIRE-001"]["item_type"] == "consumable"

    # วัสดุที่มีอยู่แล้วแต่จำนวนไม่ตรง → ต้องขึ้นเป็น update (ครุภัณฑ์ไม่เทียบจำนวน เพราะ 1 เลข = 1 ชิ้น)
    existing = {"MAT-ARD-001": {"id": 1, "name": "บอร์ด Arduino UNO R3", "location": "15310",
                                "status": "available", "item_type": "material", "quantity_total": 20}}
    diff = {d["code"]: d for d in diff_rows(rows, existing)}
    assert diff["MAT-ARD-001"]["action"] == "update"
    assert diff["MAT-ARD-001"]["changes"]["quantity_total"] == [20, 25]


def test_parse_sn_sheet(tmp_path):
    """ชีต 'SN' เสริม (2 คอลัมน์: รหัส, SN) — parse_workbook คืน sn_updates เป็น dict[code, sn]"""
    wb = openpyxl.Workbook()
    wb.active.title = "SN"
    ws = wb.active
    ws.append(["รหัส", "SN"])
    ws.append(["63-001-0001", "SN-AAA-111"])
    ws.append(["63-001-0002", "SN-BBB-222"])
    ws.append([None, "SN-ไม่มีรหัส"])  # แถวไม่มีรหัส — ต้องไม่ถูกเก็บ
    ws.append(["63-001-0003", None])   # แถวไม่มี SN — ต้องไม่ถูกเก็บ
    path = tmp_path / "sn_only.xlsx"
    wb.save(path)

    rows, skipped, sn_updates = parse_workbook(str(path))
    assert rows == []
    assert sn_updates == {"63-001-0001": "SN-AAA-111", "63-001-0002": "SN-BBB-222"}


def test_diff_rows_applies_sn_updates(tmp_path):
    """diff_rows: รหัสในชีตทะเบียนพ่วง SN เข้า changes / รหัสที่มีแต่ในชีต SN synthesize เป็น update เฉพาะ SN"""
    path = _make_register(tmp_path, [
        ("63-001-0001", "คอมพิวเตอร์ตั้งโต๊ะ Dell", "P", None, "15310"),
    ])
    rows, skipped, _sn = parse_workbook(path)

    existing = {
        "63-001-0001": {"id": 1, "name": "คอมพิวเตอร์ตั้งโต๊ะ Dell", "location": "15310",
                        "status": "available", "item_type": "durable", "quantity_total": 1,
                        "serial_number": None},
        "63-999-9999": {"id": 2, "name": "เครื่องที่ไม่อยู่ในไฟล์นี้", "location": None,
                        "status": "available", "item_type": "durable", "quantity_total": 1,
                        "serial_number": "SN-OLD"},
    }
    sn_updates = {"63-001-0001": "SN-NEW-001", "63-999-9999": "SN-NEW-999", "63-777-7777": "SN-ไม่รู้จัก"}

    diff = {d["code"]: d for d in diff_rows(rows, existing, sn_updates)}
    # รหัสที่อยู่ในชีตทะเบียนด้วย → พ่วง serial_number เข้า changes
    assert diff["63-001-0001"]["action"] == "update"
    assert diff["63-001-0001"]["changes"]["serial_number"] == [None, "SN-NEW-001"]
    # รหัสที่มีเฉพาะในชีต SN แต่พบใน DB → synthesize แถว update เฉพาะ SN
    assert diff["63-999-9999"]["action"] == "update"
    assert diff["63-999-9999"]["changes"] == {"serial_number": ["SN-OLD", "SN-NEW-999"]}
    # รหัสที่ไม่พบใน DB เลย → diff_rows ไม่สร้างแถวให้ (ผู้เรียก/preview_import ไปใส่ skipped แทน)
    assert "63-777-7777" not in diff


def test_filter_bad_sn_updates_rejects_duplicate_within_file():
    """สองรหัสในไฟล์เดียวกันอ้าง SN เดียวกัน — ต้องถูกตัดออกจาก clean ทั้งคู่ ไปที่ skipped แทน"""
    existing = {
        "A-001": {"id": 1, "name": "A", "location": None, "status": "available",
                  "item_type": "durable", "quantity_total": 1, "serial_number": None},
        "A-002": {"id": 2, "name": "B", "location": None, "status": "available",
                  "item_type": "durable", "quantity_total": 1, "serial_number": None},
    }
    sn_updates = {"A-001": "SN-DUP", "A-002": "SN-DUP"}
    clean, bad = _filter_bad_sn_updates(sn_updates, existing)
    assert clean == {}
    assert {b["seq"] for b in bad} == {"A-001", "A-002"}
    assert all("ซ้ำกับรายการอื่นในไฟล์เดียวกัน" in b["reason"] for b in bad)


def test_filter_bad_sn_updates_rejects_conflict_with_different_existing_code():
    """SN นั้นถูกใช้อยู่แล้วกับรหัสอื่นใน DB (ไม่ใช่ตัวเอง) — ต้องถูกตัดออก"""
    existing = {
        "A-001": {"id": 1, "name": "A", "location": None, "status": "available",
                  "item_type": "durable", "quantity_total": 1, "serial_number": "SN-OLD-OWNER"},
        "A-002": {"id": 2, "name": "B", "location": None, "status": "available",
                  "item_type": "durable", "quantity_total": 1, "serial_number": None},
    }
    sn_updates = {"A-002": "SN-OLD-OWNER"}
    clean, bad = _filter_bad_sn_updates(sn_updates, existing)
    assert clean == {}
    assert bad == [{"sheet": "SN", "seq": "A-002", "name": "A-002",
                    "reason": "SN นี้ถูกใช้กับรหัสอื่นในระบบแล้ว"}]


def test_filter_bad_sn_updates_allows_reassigning_same_code_its_own_sn():
    """รหัสที่แก้ SN ของตัวเองไม่ถือว่าชนกับตัวเอง (owner == code)"""
    existing = {
        "A-001": {"id": 1, "name": "A", "location": None, "status": "available",
                  "item_type": "durable", "quantity_total": 1, "serial_number": "SN-SAME"},
    }
    sn_updates = {"A-001": "SN-SAME"}
    clean, bad = _filter_bad_sn_updates(sn_updates, existing)
    assert clean == {"A-001": "SN-SAME"}
    assert bad == []


def test_reject_non_register_file(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "อย่างอื่น"
    path = tmp_path / "other.xlsx"
    wb.save(path)
    with pytest.raises(Exception):  # ไม่มีชีตที่รู้จัก → 400
        parse_workbook(str(path))
