"""นำเข้า/อัปเดตอุปกรณ์จากไฟล์ทะเบียนคุมทรัพย์สิน (Excel)

ไหลงาน 2 ขั้น เพื่อให้แอดมินตรวจก่อนของเข้าจริง:
  1. preview — อัปโหลดไฟล์ เก็บไว้ใน uploads/imports/ แล้วอ่านเทียบกับ DB คืน "ร่าง"
     ว่าจะเพิ่มอะไร แก้อะไร ชำรุด/สูญหายอะไร และอะไรหายไปจากทะเบียน (เสนอปลดระวาง)
  2. commit — อ่านไฟล์เดิมซ้ำจาก import_id แล้ว apply จริง + เขียน audit log
     (อ่านไฟล์ซ้ำแทนที่จะเชื่อ diff ที่ client ส่งกลับมา — client แก้ payload ไม่ได้)

audit log ที่เขียนตอน commit ใช้ action เดียวกับการเพิ่ม/ปลดระวางด้วยมือ
จึงออกใบรับเข้าคลัง/ใบปลดระวาง (GET /equipment/stock-document) ได้ทันทีหลังนำเข้า
"""
import os
import uuid
from typing import Any

import openpyxl
from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.config import settings
from app.models.equipment import Equipment
from app.models.equipment_category import EquipmentCategory
from app.models.user import User
from app.schemas.equipment import ImportCommitRequest
from app.services import audit_service, ocr_import

# กฎจัดหมวดตามคีย์เวิร์ด — ชุดเดียวกับ scripts/import_register.py ที่ใช้ seed ครั้งแรก
# เช็คจากบนลงล่าง เจอก่อนชนะ (คอมพิวเตอร์มาก่อนเฟอร์นิเจอร์ ไม่งั้น "คอมพิวเตอร์ตั้งโต๊ะ" เข้าหมวดโต๊ะ)
CATEGORY_RULES: list[tuple[str, list[str]]] = [
    ("คอมพิวเตอร์", ["คอมพิวเตอร์", "optiplex", "acer", "โน๊ตบุ๊ค", "โน้ตบุ๊ค",
                     "thinkpad", "แม่ข่าย", "micro pc", "กราฟฟิก", "หน้าจอ", "จอแสดงผล"]),
    ("ซอฟต์แวร์/ลิขสิทธิ์", ["ซอฟต์แวร์", "ลิขสิทธิ์", "ชุดโปรแกรม"]),
    ("โสตทัศนูปกรณ์", ["โทรทัศน์", "led tv", "smart tv"]),
    ("อุปกรณ์อิเล็กทรอนิกส์/เครื่องมือวัด",
     ["มัลติมิเตอร์", "oscilloscope", "เครื่องวัด", "เครื่องกำเนิด", "แหล่งจ่าย",
      "เครื่องทดสอบ", "ลอจิก", "iot", "ไอโอที", "sensor", "แผงวงจร", "สมองกล",
      "virtual reality", "คลัสเตอร์", "big data", "ไฟฟ้า", "เครื่องมือวัด",
      "อิเล็กทรอนิกส์", "ทดลอง", "ทดสอบ", "อุปกรณ์วัด"]),
    ("โต๊ะ/ตู้/เฟอร์นิเจอร์", ["โต๊ะ", "โตะ", "เก้าอี้", "ตู้", "ชั้นวาง"]),
]
FALLBACK_CAT = "ไม่ระบุหมวดหมู่"

# รหัสวัสดุ 2 หลักแรก = หมวดวัสดุ (ครุภัณฑ์ไม่ใช้ — 2 หลักแรกของครุภัณฑ์คือปีงบประมาณ 63-67)
MATERIAL_CAT_BY_PREFIX: dict[str, str] = {
    "10": "อุปกรณ์ใช้งานทั่วไป",
    "12": "อุปกรณ์คอมพิวเตอร์/ต่อพ่วง",   # webcam
    "13": "อุปกรณ์คอมพิวเตอร์/ต่อพ่วง",   # converter, hub, card reader
    "21": "บอร์ด/ไมโครคอนโทรลเลอร์",
    "22": "บอร์ด/ไมโครคอนโทรลเลอร์",
    "23": "บอร์ด/ไมโครคอนโทรลเลอร์",
    "30": "เครื่องมือช่าง",
    "33": "ชุดการเรียนรู้",
    "34": "อุปกรณ์อิเล็กทรอนิกส์/เครื่องมือวัด",
    "35": "อุปกรณ์อิเล็กทรอนิกส์/เครื่องมือวัด",
    "36": "อุปกรณ์อิเล็กทรอนิกส์/เครื่องมือวัด",
    "61": "อุปกรณ์เครือข่าย",
}

MAX_UPLOAD_BYTES = 10 * 1024 * 1024

# ประเภทในคอลัมน์ "ประเภท" ของชีตวัสดุ → item_type ในระบบ (ไม่ตรงคำไหน = วัสดุใช้ซ้ำ)
_MATERIAL_TYPES = {
    "วัสดุใช้ซ้ำ": "material", "material": "material",
    "วัสดุสิ้นเปลือง": "consumable", "consumable": "consumable",
}


def _clean(v: object) -> str:
    """ตัด whitespace/ขึ้นบรรทัดที่ติดมาจากเซลล์ Excel"""
    return " ".join(str(v).split()) if v is not None else ""


def _categorize(name: str, code: str = "", item_type: str = "durable") -> str:
    """หมวดหมู่ของรายการ — วัสดุใช้รหัส 2 หลักแรก (หมวดวัสดุ) ก่อน ไม่เจอค่อยเดาจากชื่อ"""
    if item_type != "durable" and code[:2] in MATERIAL_CAT_BY_PREFIX:
        return MATERIAL_CAT_BY_PREFIX[code[:2]]
    low = name.lower()
    for cat, keys in CATEGORY_RULES:
        if any(k in low for k in keys):
            return cat
    return FALLBACK_CAT


def _status_from(normal: object, broken: object, worn: object, lost: object) -> tuple[str, int]:
    """แปลงคอลัมน์ ปกติ/ชำรุด/เสื่อมสภาพ/สูญหาย ในทะเบียน → (status, quantity_available)"""
    if broken:
        return "damaged", 0
    if lost or worn:
        return "unavailable", 0
    return "available", 1


def parse_workbook(path: str) -> tuple[list[dict[str, Any]], list[dict[str, str]], dict[str, str]]:
    """อ่านไฟล์ทะเบียน คืน (รายการที่มีเลขครุภัณฑ์, รายการที่ไม่มีเลข — ข้ามไป, SN จากชีตเสริม "SN")

    รองรับ 2 ชีตของทะเบียนคณะ: ชีตคณะ (เลขปกติ) และชีตพระราชทาน (ใช้ "รหัสครุภัณฑ์ใหม่")
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="อ่านไฟล์ Excel ไม่สำเร็จ")

    rows: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []

    def add(code: str, name: str, loc: str, st: str, qa: int, sheet: str, seq: str,
            item_type: str = "durable", qty: int = 1, unit: str | None = None) -> None:
        if not code or code == "-":
            skipped.append({"sheet": sheet, "seq": seq, "name": name, "reason": "ไม่มีเลขครุภัณฑ์/รหัส"})
            return
        rows.append({
            "code": code, "name": name, "location": loc or None,
            "status": st, "quantity_available": qa, "category": _categorize(name, code, item_type),
            "item_type": item_type, "quantity_total": qty, "unit": unit,
        })

    if "คณะเทคโนฯดิจิทัล" in wb.sheetnames:
        for r in wb["คณะเทคโนฯดิจิทัล"].iter_rows(min_row=5, values_only=True):
            name = _clean(r[2])
            if not name or name == "รวมทั้งสิ้น":
                continue
            st, qa = _status_from(r[5], r[6], r[7], r[8])
            add(_clean(r[1]), name, _clean(r[11]) or _clean(r[16]), st, qa, "คณะ", _clean(r[0]))

    if "ครุภัณฑ์ที่ได้รับพระราชทาน" in wb.sheetnames:
        for r in wb["ครุภัณฑ์ที่ได้รับพระราชทาน"].iter_rows(min_row=5, values_only=True):
            name = _clean(r[1])
            if not name or name == "รวมทั้งสิ้น":
                continue
            st, qa = _status_from(r[10], r[11], r[12], r[13])
            # "ไม่พบครุภัณฑ์" ในหมายเหตุ = หาตัวจริงไม่เจอ → ยืมไม่ได้
            if "ไม่พบครุภัณฑ์" in _clean(r[15]):
                st, qa = "unavailable", 0
            add(_clean(r[8]), name, _clean(r[9]), st, qa, "พระราชทาน", _clean(r[0]))

    # ชีต "วัสดุ" (ทางเลือก) — วัสดุไม่มีในทะเบียนคุมทรัพย์สิน จึงใช้ชีตแยกรูปแบบง่าย ๆ
    # header แถว 1: รหัส | รายการ | ประเภท | หน่วยนับ | จำนวน | สถานที่จัดเก็บ
    if "วัสดุ" in wb.sheetnames:
        for i, r in enumerate(wb["วัสดุ"].iter_rows(min_row=2, values_only=True), 2):
            name = _clean(r[1]) if len(r) > 1 else ""
            if not name:
                continue
            item_type = _MATERIAL_TYPES.get(_clean(r[2]), "material")
            qty = int(r[4] or 0) if len(r) > 4 else 0
            add(_clean(r[0]), name, _clean(r[5]) if len(r) > 5 else "",
                "available" if qty > 0 else "unavailable", qty, "วัสดุ", str(i),
                item_type=item_type, qty=qty, unit=_clean(r[3]) or None)

    # ชีต "SN" (ทางเลือก) — SN จากผู้ผลิต ไม่มีในทะเบียนคุมทรัพย์สิน/ชีตวัสดุ จึงเสริมมาแยกต่างหาก
    # header แถว 1: รหัส | SN — update-only จับคู่ด้วยรหัสที่มีอยู่แล้วในระบบเท่านั้น (ดู diff_rows)
    sn_updates: dict[str, str] = {}
    if "SN" in wb.sheetnames:
        for r in wb["SN"].iter_rows(min_row=2, values_only=True):
            code = _clean(r[0]) if len(r) > 0 else ""
            sn = _clean(r[1]) if len(r) > 1 else ""
            if not code or not sn:
                continue
            sn_updates[code] = sn

    wb.close()
    # ไฟล์ที่มีแต่ชีต "SN" ล้วน (แยกไฟล์จากทะเบียนหลัก) ต้องผ่านได้ ไม่ใช่แค่ไฟล์ทะเบียน/วัสดุเท่านั้น
    if not rows and not skipped and not sn_updates:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ไม่พบข้อมูลในไฟล์ — ต้องมีชีต 'คณะเทคโนฯดิจิทัล', 'ครุภัณฑ์ที่ได้รับพระราชทาน', 'วัสดุ' หรือ 'SN'",
        )
    return rows, skipped, sn_updates


STATUS_LABEL = {"available": "ปกติ", "damaged": "ชำรุด", "unavailable": "สูญหาย/เสื่อมสภาพ",
                "retired": "ปลดระวาง", "under_repair": "ซ่อม", "borrowed": "ถูกยืม"}


def diff_rows(
    rows: list[dict[str, Any]],
    existing: dict[str, dict[str, Any]],
    sn_updates: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """เทียบรายการในไฟล์กับของใน DB → คืนร่างการเปลี่ยนแปลงทีละรายการ

    action:
      new       = ยังไม่มีใน DB → จะเพิ่มเข้าคลัง
      update    = มีอยู่แล้วแต่ชื่อ/สถานที่/สถานะ/SN ไม่ตรง → จะแก้ตามไฟล์
      unchanged = ตรงกันทุกอย่าง → ไม่ทำอะไร
      retire    = มีใน DB (ครุภัณฑ์) แต่หายไปจากไฟล์ทะเบียน → เสนอปลดระวาง (ต้องติ๊กยืนยันแยก)

    sn_updates: รหัส → SN จากชีตเสริม "SN" (update-only จับคู่กับรหัสที่มีอยู่แล้วในระบบเท่านั้น ไม่สร้างของใหม่)
      - รหัสที่อยู่ในชีตทะเบียน/วัสดุของไฟล์นี้ด้วย → พ่วง serial_number เข้า changes เหมือนฟิลด์อื่น
      - รหัสที่มีเฉพาะในชีต SN แต่พบใน DB → synthesize แถว action=update เฉพาะ SN (ฟิลด์อื่นคงค่าเดิม)
      - รหัสที่ไม่พบใน DB เลย → ผู้เรียก (preview_import) เอาไปใส่ใน skipped แทน (เหตุผล "ไม่พบรหัสนี้ในระบบ")
    """
    sn_updates = sn_updates or {}
    result: list[dict[str, Any]] = []
    in_file = {row["code"] for row in rows}

    for row in rows:
        old = existing.get(row["code"])
        if old is None:
            # รหัสใหม่ที่ยังไม่มีใน DB แต่ก็มีอยู่ในชีต SN ของไฟล์เดียวกันด้วย — พ่วง SN เข้าไปให้แอดมิน
            # เห็นในร่างตั้งแต่ตอน preview (ไม่งั้นจะถูกสร้างพร้อม SN จริงตอน commit โดยที่ไม่เคยผ่านสายตา
            # แอดมินมาก่อนเลย ผิดหลักการของฟีเจอร์นี้ทั้งหมดที่ต้องให้แอดมินตรวจ/แก้ร่างก่อนของเข้าจริง)
            sn = sn_updates.get(row["code"])
            changes = {"serial_number": [None, sn]} if sn else {}
            result.append({"action": "new", **row, "serial_number": sn, "changes": changes})
            continue
        changes: dict[str, list[Any]] = {}
        for field in ("name", "location", "status"):
            if (old.get(field) or None) != (row.get(field) or None):
                changes[field] = [old.get(field), row.get(field)]
        # จำนวนเทียบเฉพาะวัสดุ — ครุภัณฑ์เป็น 1 ชิ้น/1 เลขทะเบียนเสมอ
        if row["item_type"] != "durable" and old.get("quantity_total") != row["quantity_total"]:
            changes["quantity_total"] = [old.get("quantity_total"), row["quantity_total"]]
        sn = sn_updates.get(row["code"])
        target_sn = sn or old.get("serial_number")
        if sn and (old.get("serial_number") or None) != sn:
            changes["serial_number"] = [old.get("serial_number"), sn]
        result.append({
            "action": "update" if changes else "unchanged", **row,
            "item_type": old["item_type"],  # ประเภทของเดิมใน DB ชนะ (แอดมินอาจจัดประเภทไว้แล้ว)
            "serial_number": target_sn,
            "changes": changes,
        })

    # เสนอปลดระวางได้ก็ต่อเมื่อไฟล์นี้มีรายการทะเบียน/วัสดุจริง (rows ไม่ว่าง) — ไฟล์ที่มีแต่ชีต "SN" ล้วน
    # (rows ว่าง) ไม่ใช่ทะเบียนฉบับเต็ม ถ้าไม่กันไว้ตรงนี้จะเข้าใจผิดว่าครุภัณฑ์ทุกชิ้นในระบบหายไปจากไฟล์
    # แล้วเสนอปลดระวางทั้งดุ้น ทั้งที่ไฟล์นี้ไม่ได้พูดถึงทะเบียนเลย
    if rows:
        for code, old in existing.items():
            # เฉพาะครุภัณฑ์ที่ยังไม่ปลดระวาง — วัสดุ (material/consumable) ไม่ได้อยู่ในทะเบียนอยู่แล้ว
            if code not in in_file and old["item_type"] == "durable" and old["status"] != "retired":
                result.append({
                    "action": "retire", "code": code, "name": old["name"],
                    "location": old.get("location"), "status": "retired",
                    "quantity_available": 0, "category": None, "item_type": old["item_type"],
                    "quantity_total": old.get("quantity_total", 1), "unit": None, "changes": {},
                })

    # รหัสที่มีเฉพาะในชีต SN (ไม่อยู่ในชีตทะเบียน/วัสดุของไฟล์นี้) แต่จับคู่กับของที่มีอยู่แล้วใน DB ได้
    # → synthesize แถว action=update เฉพาะ SN ฟิลด์อื่นคงค่าเดิมทุกอย่าง (update-only ตามที่ตั้งใจ)
    for code, sn in sn_updates.items():
        if code in in_file:
            continue
        old = existing.get(code)
        if old is None:
            continue  # ไม่พบรหัสนี้ในระบบเลย — preview_import เอาไปใส่ skipped แทน
        if (old.get("serial_number") or None) == sn:
            continue  # ตรงกับของเดิมอยู่แล้ว ไม่มีอะไรต้องเปลี่ยน
        result.append({
            "action": "update", "code": code, "name": old["name"],
            "location": old.get("location"), "status": old["status"],
            "quantity_available": 0, "category": None, "item_type": old["item_type"],
            "quantity_total": old.get("quantity_total", 1), "unit": None,
            "serial_number": sn,
            "changes": {"serial_number": [old.get("serial_number"), sn]},
        })
    return result


def _filter_bad_sn_updates(
    sn_updates: dict[str, str], existing: dict[str, dict[str, Any]]
) -> tuple[dict[str, str], list[dict[str, str]]]:
    """คัดกรอง SN ที่จะชนกันออกจาก sn_updates ก่อนส่งเข้า diff_rows/commit — กัน UniqueViolationError
    บน partial unique index (ix_equipment_serial_number_unique) ตอน commit ซึ่งไม่มี IntegrityError
    handler ที่ไหนใน backend เลย จึงกลายเป็น 500 ที่ยกเลิก commit ทั้งดุ้นหลังแอดมินตรวจร่างผ่านแล้ว

    ตัดออก 2 กรณี (คืนไปเป็น skipped แทน ไม่ทำให้ทั้งบรรทัดที่แตะรหัสนั้นพังไปด้วย — แค่ไม่ apply ค่า SN นี้):
      (a) SN ค่าเดียวกันถูกใช้ซ้ำโดยมากกว่า 1 รหัสในชีต SN ของไฟล์เดียวกันเอง
      (b) SN ค่านั้นถูกใช้อยู่แล้วกับรหัสอื่นที่ไม่ใช่ตัวมันเองใน DB
    """
    value_counts: dict[str, int] = {}
    for sn in sn_updates.values():
        value_counts[sn] = value_counts.get(sn, 0) + 1

    existing_sn_owner = {
        eq["serial_number"]: code for code, eq in existing.items() if eq.get("serial_number")
    }

    clean: dict[str, str] = {}
    bad: list[dict[str, str]] = []
    for code, sn in sn_updates.items():
        if value_counts[sn] > 1:
            bad.append({"sheet": "SN", "seq": code, "name": code,
                        "reason": "SN ซ้ำกับรายการอื่นในไฟล์เดียวกัน"})
            continue
        owner = existing_sn_owner.get(sn)
        if owner is not None and owner != code:
            bad.append({"sheet": "SN", "seq": code, "name": code,
                        "reason": "SN นี้ถูกใช้กับรหัสอื่นในระบบแล้ว"})
            continue
        clean[code] = sn
    return clean, bad


async def _load_existing(db: AsyncSession) -> dict[str, dict[str, Any]]:
    result = await db.execute(select(Equipment))
    return {
        eq.code: {"id": eq.id, "name": eq.name, "location": eq.location, "status": eq.status,
                  "item_type": eq.item_type, "quantity_total": eq.quantity_total,
                  "serial_number": eq.serial_number}
        for eq in result.scalars().all()
    }


def _summarize(diff: list[dict[str, Any]]) -> dict[str, int]:
    counts = {"new": 0, "update": 0, "unchanged": 0, "retire": 0}
    for d in diff:
        counts[d["action"]] += 1
    return counts


def _parse_file(path: str) -> tuple[list[dict[str, Any]], list[dict[str, str]], dict[str, str]]:
    """Excel → อ่านตามชีตทะเบียน / รูป-PDF → OCR (โหมดทดลอง) — คืน rows รูปแบบเดียวกัน

    OCR ไม่รองรับชีต SN (อ่านจากรูป/PDF เดียว ไม่มีแนวคิดหลายชีต) จึงคืน sn_updates ว่างเสมอ
    """
    if ocr_import.is_scan(path):
        rows, skipped = ocr_import.parse_scan(path)
        return rows, skipped, {}
    return parse_workbook(path)


def _import_path(import_id: str) -> str | None:
    """หาไฟล์ที่อัปโหลดไว้จาก import_id (นามสกุลต่างกันได้ตามชนิดไฟล์ที่แอดมินอัปมา)"""
    for ext in (".xlsx", *ocr_import.SCAN_EXTS):
        path = os.path.join(settings.IMPORT_DIR, f"{import_id}{ext}")
        if os.path.exists(path):
            return path
    return None


async def preview_import(db: AsyncSession, file: UploadFile) -> dict[str, Any]:
    """รับไฟล์ทะเบียน (Excel หรือรูป/PDF) เก็บไว้ แล้วคืนร่างว่าจะเปลี่ยนอะไรบ้าง (ยังไม่แตะ DB)"""
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext != ".xlsx" and not ocr_import.is_scan(file.filename or ""):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="รองรับไฟล์ .xlsx หรือรูปถ่าย/PDF ของทะเบียน (.jpg .png .pdf)")
    # เช็คขนาดก่อน read() — ไม่งั้นไฟล์ใหญ่เท่าไรก็ถูกโหลดเข้า RAM หมดก่อนถึงบรรทัดตรวจ
    if file.size is not None and file.size > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ไฟล์ใหญ่เกิน 10MB")
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:  # เผื่อกรณี .size เป็น None
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ไฟล์ใหญ่เกิน 10MB")

    # เก็บนอก UPLOAD_DIR — ที่นั่นถูก mount เป็น StaticFiles สาธารณะ ไฟล์ทะเบียนคณะจะหลุด
    os.makedirs(settings.IMPORT_DIR, exist_ok=True)
    import_id = uuid.uuid4().hex
    path = os.path.join(settings.IMPORT_DIR, f"{import_id}{ext}")
    with open(path, "wb") as f:
        f.write(contents)

    rows, skipped, sn_updates = _parse_file(path)
    existing = await _load_existing(db)
    sn_updates, sn_bad = _filter_bad_sn_updates(sn_updates, existing)
    skipped.extend(sn_bad)
    diff = diff_rows(rows, existing, sn_updates)
    # รหัสที่อยู่เฉพาะในชีต SN แต่ไม่พบใน DB เลย — update-only ไม่สร้างของใหม่ ไปโผล่เป็น "ข้าม" แทน
    in_file = {row["code"] for row in rows}
    for code in sn_updates:
        if code not in in_file and code not in existing:
            skipped.append({"sheet": "SN", "seq": code, "name": code, "reason": "ไม่พบรหัสนี้ในระบบ"})
    return {
        "import_id": import_id,
        "filename": file.filename,
        "summary": _summarize(diff),
        "rows": diff,
        "skipped": skipped,
    }


async def _get_or_create_category(db: AsyncSession, name: str, cache: dict[str, EquipmentCategory]) -> EquipmentCategory:
    if name in cache:
        return cache[name]
    cat = (await db.execute(
        select(EquipmentCategory).where(EquipmentCategory.name == name)
    )).scalar_one_or_none()
    if cat is None:
        cat = EquipmentCategory(name=name)
        db.add(cat)
        await db.flush()
    cache[name] = cat
    return cat


async def _assert_no_sn_conflict(db: AsyncSession, sn: str, exclude_id: uuid.UUID | None) -> None:
    """เช็คก่อนเขียน SN ว่าไม่ชนกับรหัสอื่นใน DB — defense-in-depth เผื่อ DB เปลี่ยนระหว่าง preview→commit
    (เช่นมีคนแก้ไข SN ผ่านช่องทางอื่นแทรกมา) หรือ client เลี่ยง validation ตอน preview แล้วยัด SN มาตรง ๆ
    ใน body — กัน asyncpg.UniqueViolationError บน ix_equipment_serial_number_unique กลายเป็น 500 ที่ไม่มี
    IntegrityError handler ไหนใน backend ดักไว้เลย ให้เป็น 409 อ่านง่ายแทน
    """
    query = select(Equipment.code).where(Equipment.serial_number == sn)
    if exclude_id is not None:
        query = query.where(Equipment.id != exclude_id)
    dup_code = (await db.execute(query)).scalar_one_or_none()
    if dup_code is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"SN '{sn}' ถูกใช้กับรหัส {dup_code} อยู่แล้ว",
        )


async def commit_import(
    db: AsyncSession, admin: User, import_id: str, body: ImportCommitRequest
) -> dict[str, int]:
    """บันทึกร่างที่แอดมินตรวจ/แก้แล้วเข้าระบบจริง — เพิ่ม / แก้ / ปลดระวาง ตามบรรทัดที่ส่งมา

    บรรทัดที่แอดมินตัดออกจากร่าง จะไม่ถูกบันทึก และไฟล์ต้นทางถูกอ่านซ้ำเพื่อยืนยันว่า
    ทุกบรรทัดที่ส่งมามีอยู่ในไฟล์จริงและ action ตรงกับที่ระบบคำนวณไว้ (กันร่างถูกยัดของแปลกปลอม)
    แต่ค่า ชื่อ/สถานที่/สถานะ ใช้ตามที่แอดมินแก้ในร่าง

    เขียน audit log ทุกบรรทัด ทำให้ออกใบรับเข้าคลัง/ใบปลดระวางตามช่วงวันที่ได้ทันทีหลังบันทึก
    """
    if not import_id.isalnum():  # กัน path traversal จาก import_id ที่ปลอมมา
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="import_id ไม่ถูกต้อง")
    path = _import_path(import_id)
    if not path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="ไม่พบไฟล์นำเข้า กรุณาอัปโหลดใหม่")

    rows, _, sn_updates = _parse_file(path)
    existing = await _load_existing(db)
    sn_updates, _sn_bad = _filter_bad_sn_updates(sn_updates, existing)
    allowed = {d["code"]: d for d in diff_rows(rows, existing, sn_updates)}

    applied = {"new": 0, "update": 0, "retire": 0}
    cat_cache: dict[str, EquipmentCategory] = {}
    src = body.filename or os.path.basename(path)

    for row in body.rows:
        ref = allowed.get(row.code)
        if ref is None or ref["action"] != row.action or row.action == "unchanged":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"รายการ {row.code} ไม่ตรงกับไฟล์ที่อัปโหลด กรุณาอ่านไฟล์ใหม่",
            )

        cat_names = row.categories or [ref.get("category") or FALLBACK_CAT]

        if row.action == "new":
            # SN ของรหัสใหม่ตอนนี้มาจาก diff_rows แล้ว (พ่วงเข้า row ตั้งแต่ตอน preview — ดู comment ใน
            # diff_rows) ดังนั้น row.serial_number ที่แอดมินเห็น/แก้ในร่างคือค่าที่ใช้ apply จริงตรง ๆ
            # ไม่ fallback ไป sn_updates เองอีกชั้น เพื่อให้ค่าที่บันทึกตรงกับที่แอดมินตรวจแล้วเป๊ะ
            sn = row.serial_number
            if sn:
                await _assert_no_sn_conflict(db, sn, None)
            eq = Equipment(
                code=row.code, name=row.name, item_type=row.item_type,
                location=row.location, status=row.status,
                unit=row.unit, description=row.description, serial_number=sn,
                quantity_total=row.quantity,
                quantity_available=row.quantity if row.status == "available" else 0,
                image_urls=row.image_urls, image_url=row.image_urls[0] if row.image_urls else None,
            )
            eq.categories = [await _get_or_create_category(db, n, cat_cache) for n in cat_names]
            db.add(eq)
            await db.flush()
            await audit_service.log_action(
                db, admin, "create_equipment", "equipment", eq.id,
                {"code": eq.code, "name": eq.name, "quantity": eq.quantity_total,
                 "item_type": eq.item_type, "source": src},
            )
            applied["new"] += 1
            continue

        eq = (await db.execute(
            select(Equipment).where(Equipment.id == existing[row.code]["id"]).options(selectinload(Equipment.categories))
        )).scalar_one()

        if row.action == "retire":
            eq.status = "retired"
            eq.quantity_available = 0
            await audit_service.log_action(
                db, admin, "retire_equipment", "equipment", eq.id,
                {"code": eq.code, "name": eq.name, "quantity": eq.quantity_total,
                 "item_type": eq.item_type,
                 "reason": (row.reason or "").strip() or f"ไม่พบในทะเบียนฉบับ {src}"},
            )
            applied["retire"] += 1
            continue

        # action == update — เขียนทับด้วยค่าที่แอดมินยืนยันในร่าง
        old_status, old_qty = eq.status, eq.quantity_total
        eq.name, eq.location, eq.status = row.name, row.location, row.status
        eq.item_type, eq.quantity_total = row.item_type, row.quantity
        if row.unit is not None:
            eq.unit = row.unit
        if row.description is not None:
            eq.description = row.description
        # แตะเฉพาะแถวที่ร่างมีค่า SN ให้ — กันแถวที่ไม่ได้แตะ SN ไปเขียนทับ SN เดิมเป็น null
        if row.serial_number is not None:
            if row.serial_number != eq.serial_number:
                await _assert_no_sn_conflict(db, row.serial_number, eq.id)
            eq.serial_number = row.serial_number
        if row.image_urls:
            eq.image_urls = row.image_urls
            eq.image_url = row.image_urls[0]
        if row.categories:
            eq.categories = [await _get_or_create_category(db, n, cat_cache) for n in row.categories]
        # แตะ quantity_available เฉพาะตอนสถานะ/จำนวนรวมเปลี่ยน — ขยับตามส่วนต่าง (delta) ไม่ใช่ตั้งค่าทับ
        # ตรง ๆ เพราะการตั้งทับ (= row.quantity หรือ 0) จะทิ้งจำนวนที่ถูกยืมออกไปจริงในตอนนั้นไปเลย (เจอบั๊กจริง
        # — นำเข้าซ้ำแล้วค่าคงเหลือเพี้ยนค้างถาวร ไม่มีช่องแก้ quantity_available ตรง ๆ ที่ไหนให้แก้คืน)
        if row.status != old_status or row.quantity != old_qty:
            if row.status == "available":
                eq.quantity_available = max(0, min(row.quantity, eq.quantity_available + (row.quantity - old_qty)))
            else:
                eq.quantity_available = 0
        await audit_service.log_action(
            db, admin, "update_equipment", "equipment", eq.id,
            {"code": eq.code, "changes": ref["changes"], "source": src},
        )
        applied["update"] += 1

    await db.commit()
    return applied
